# One-off script to dump Excel structure for dev
import json
import openpyxl
wb = openpyxl.load_workbook("domino_inventory_training.xlsx", read_only=True, data_only=True)
out = {}
for name in wb.sheetnames:
    ws = wb[name]
    rows = []
    for r in ws.iter_rows(values_only=True):
        rows.append([str(c).strip() if c is not None else "" for c in r])
    out[name] = rows
wb.close()
with open("xlsx_dump.json", "w", encoding="utf-8") as f:
    json.dump(out, f, ensure_ascii=False, indent=2)
print("Written xlsx_dump.json")
